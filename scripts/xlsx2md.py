#!/usr/bin/env python3
"""Convert Excel (.xlsx) spreadsheets to Markdown.

Features:
  - Each worksheet -> a Markdown section with sheet name as heading
  - Tables -> HTML tables with rowspan/colspan for merged cells
  - Inline formatting: bold, italic detection
  - Multi-line cell content support
  - Empty row/column trimming

Usage:
  python xlsx2md.py input.xlsx [-o output.md]

Can also be imported as a module:
  from xlsx2md import convert
  convert("input.xlsx", "output.md")

Requires: openpyxl (pip install openpyxl)
"""

import sys
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl",
          file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Inline formatting helpers
# ---------------------------------------------------------------------------

def _escape_md(text: str) -> str:
    """Escape Markdown special characters in plain text."""
    if not text:
        return ""
    for ch in (r"\\", "`", "*", "_", "[", "]", "#"):
        text = text.replace(ch, "\\" + ch)
    return text


def _cell_inline_html(text: str, font) -> str:
    """Apply inline formatting to cell text based on font properties."""
    if not text:
        return ""

    escaped = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    # Handle newlines within cells
    escaped = escaped.replace("\n", "<br>")

    if font:
        is_bold = font.bold
        is_italic = font.italic

        if is_bold and is_italic:
            return f"<strong><em>{escaped}</em></strong>"
        elif is_bold:
            return f"<strong>{escaped}</strong>"
        elif is_italic:
            return f"<em>{escaped}</em>"

    return escaped


# ---------------------------------------------------------------------------
# Merged cell map builder
# ---------------------------------------------------------------------------

def _build_merge_map(ws) -> dict:
    """Build a map of (row, col) -> (rowspan, colspan, is_continuation).

    For the top-left cell of a merged range: (rowspan, colspan, False)
    For continuation cells: (0, 0, True) — should be skipped.
    """
    merge_map = {}

    for merge_range in ws.merged_cells.ranges:
        min_row = merge_range.min_row
        max_row = merge_range.max_row
        min_col = merge_range.min_col
        max_col = merge_range.max_col

        rowspan = max_row - min_row + 1
        colspan = max_col - min_col + 1

        # Top-left cell gets the span info
        merge_map[(min_row, min_col)] = (rowspan, colspan, False)

        # All other cells in the range are continuations
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) != (min_row, min_col):
                    merge_map[(r, c)] = (0, 0, True)

    return merge_map


# ---------------------------------------------------------------------------
# Sheet -> HTML table
# ---------------------------------------------------------------------------

def _sheet_to_html(ws) -> str:
    """Convert a worksheet to an HTML table with merged cell support."""
    merge_map = _build_merge_map(ws)

    # Determine the actual data range (trim empty rows/cols at edges)
    min_row = ws.min_row
    max_row = ws.max_row
    min_col = ws.min_column
    max_col = ws.max_column

    if min_row is None or max_row is None:
        return ""
    if min_col is None or max_col is None:
        return ""

    # Trim trailing empty rows
    while max_row > min_row:
        row_has_data = False
        for c in range(min_col, max_col + 1):
            if (max_row, c) in merge_map:
                rs, cs, is_cont = merge_map[(max_row, c)]
                if not is_cont:
                    row_has_data = True
                    break
            elif ws.cell(row=max_row, column=c).value is not None:
                row_has_data = True
                break
        if row_has_data:
            break
        max_row -= 1

    # Trim trailing empty columns
    while max_col > min_col:
        col_has_data = False
        for r in range(min_row, max_row + 1):
            if (r, max_col) in merge_map:
                rs, cs, is_cont = merge_map[(r, max_col)]
                if not is_cont:
                    col_has_data = True
                    break
            elif ws.cell(row=r, column=max_col).value is not None:
                col_has_data = True
                break
        if col_has_data:
            break
        max_col -= 1

    n_rows = max_row - min_row + 1
    n_cols = max_col - min_col + 1

    if n_rows <= 0 or n_cols <= 0:
        return ""

    # Track which positions are covered by a span
    covered = set()

    html = ["<table>"]

    for r in range(min_row, max_row + 1):
        html.append("  <tr>")

        for c in range(min_col, max_col + 1):
            # Skip if covered by a previous span
            if (r, c) in covered:
                continue

            # Check if this cell is part of a merge
            if (r, c) in merge_map:
                rs, cs, is_cont = merge_map[(r, c)]
                if is_cont:
                    # Continuation cell — should have been covered already
                    continue
                # Leading cell of a merged range
                for dr in range(rs):
                    for dc in range(cs):
                        if dr == 0 and dc == 0:
                            continue
                        covered.add((r + dr, c + dc))
            else:
                rs, cs = 1, 1

            # Get cell value and formatting
            cell = ws.cell(row=r, column=c)
            raw_value = cell.value
            if raw_value is None:
                raw_value = ""
            text = str(raw_value).strip()

            # Apply inline formatting
            formatted = _cell_inline_html(text, cell.font)

            # Build cell HTML
            attrs = ""
            if cs > 1:
                attrs += f' colspan="{cs}"'
            if rs > 1:
                attrs += f' rowspan="{rs}"'

            tag = "th" if r == min_row else "td"
            html.append(f"    <{tag}{attrs}>{formatted}</{tag}>")

        html.append("  </tr>")

    html.append("</table>")
    return "\n".join(html)


# ---------------------------------------------------------------------------
# Main conversion
# ---------------------------------------------------------------------------

def convert(xlsx_path: str, output_path: str | None = None) -> str:
    """Convert an .xlsx file to Markdown and return the Markdown string."""
    wb = load_workbook(xlsx_path, data_only=True)

    sections = []
    total_sheets = len(wb.sheetnames)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        html = _sheet_to_html(ws)

        if not html:
            continue

        parts = []
        # Use sheet name as heading
        if total_sheets > 1:
            parts.append(f"## {sheet_name}")
        parts.append(html)
        sections.append("\n\n".join(parts))

    result = "\n\n---\n\n".join(sections)

    # Clean up excessive blank lines
    while "\n\n\n\n" in result:
        result = result.replace("\n\n\n\n", "\n\n\n")
    while "\n\n\n" in result:
        result = result.replace("\n\n\n", "\n\n")
    result = result.strip() + "\n"

    wb.close()

    if output_path:
        Path(output_path).write_text(result, encoding="utf-8")
        print(f"OK: {xlsx_path} -> {output_path}  ({total_sheets} sheets)",
              file=sys.stderr)

    return result


def main():
    ap = argparse.ArgumentParser(
        description="Convert Excel (.xlsx) to Markdown",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("input", help="Input .xlsx file path")
    ap.add_argument("-o", "--output", help="Output .md file (default: <input>.md)")
    args = ap.parse_args()

    inp = Path(args.input)
    if not inp.exists():
        print(f"Error: file not found: {inp}", file=sys.stderr)
        sys.exit(1)

    out = args.output or str(inp.with_suffix(".md"))
    convert(str(inp), out)


if __name__ == "__main__":
    main()
